from openpyxl import load_workbook
import numpy as np
import pandas as pd

filename = './data/2022 Kestane Alanları.xlsx'

# Load the Excel file
wb = load_workbook(filename)

# Get the sheet name
sheet_name = wb.sheetnames[0]

# Load the sheet as a dataframe
df = pd.read_excel(filename, sheet_name=sheet_name)

data = np.ndarray([], np.uint32)

for i in range(len(df)):
   data = np.append(data, df.loc[i])

data = data.tolist()
data = [data[i:i+4] for i in range(1, len(data), 4)]
print(data)

with open('data.txt', 'w+', encoding="utf-8") as f:
   for i in data:
      string = str(i[0]) +","+ str(i[1]) +","+ str(i[2]) +","+ str(i[3])
      f.writelines(string)
      f.write("\n")